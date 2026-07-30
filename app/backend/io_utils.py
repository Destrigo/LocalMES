"""Spreadsheet / CSV helpers for LocalMES imports and exports."""

from __future__ import annotations

import csv
import io
from typing import Iterable

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from fastapi import HTTPException, UploadFile
from fastapi.responses import StreamingResponse


async def read_upload_bytes(file: UploadFile) -> bytes:
    data = await file.read()
    if not data:
        raise HTTPException(400, "Empty file")
    return data


def xlsx_sheet_to_dicts(content: bytes, sheet_name: str | None = None) -> list[dict]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(400, f"Invalid Excel file: {exc}") from exc
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise HTTPException(400, f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower().replace(" ", "_") if h is not None else "" for h in rows[0]]
    out: list[dict] = []
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        item = {}
        for i, v in enumerate(row):
            if i < len(headers) and headers[i]:
                item[headers[i]] = "" if v is None else str(v).strip()
        out.append(item)
    return out


def csv_to_dicts(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    sample = text[:8192]
    try:
        delim = csv.Sniffer().sniff(sample, delimiters=";,").delimiter
    except csv.Error:
        delim = "," if sample.count(",") >= sample.count(";") else ";"
    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    rows = []
    for raw in reader:
        rows.append(
            {
                (k or "").strip().lower().replace(" ", "_"): (v or "").strip()
                for k, v in raw.items()
                if k
            }
        )
    return rows


async def parse_tabular(file: UploadFile, sheet_name: str | None = None) -> list[dict]:
    content = await read_upload_bytes(file)
    name = (file.filename or "").lower()
    if name.endswith(".csv"):
        return csv_to_dicts(content)
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return xlsx_sheet_to_dicts(content, sheet_name=sheet_name)
    # try excel then csv
    try:
        return xlsx_sheet_to_dicts(content, sheet_name=sheet_name)
    except HTTPException:
        return csv_to_dicts(content)


def xlsx_response(
    sheets: dict[str, tuple[list[str], list[list]]],
    filename: str,
) -> StreamingResponse:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(bold=True, color="FFFFFF")
    for title, (headers, rows) in sheets.items():
        ws = wb.create_sheet(title[:31])
        for col, val in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=val)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def pdf_table_response(
    title: str,
    headers: list[str],
    rows: Iterable[list],
    filename: str = "report.pdf",
) -> StreamingResponse:
    from datetime import datetime

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = io.BytesIO()
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1 * cm,
        rightMargin=1 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    story = [
        Paragraph(f"<b>{title}</b>", styles["Title"]),
        Paragraph(datetime.utcnow().strftime("Generated %Y-%m-%d %H:%M UTC"), styles["Normal"]),
        Spacer(1, 0.5 * cm),
    ]
    data = [headers] + [list(r) for r in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
